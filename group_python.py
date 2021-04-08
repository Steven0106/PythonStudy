from openpyxl import load_workbook
from openpyxl import Workbook
from glob import *
import os

# 기존 파일이 있으면 삭제해주는 역할
try:
    os.remove("group_python.xlsx")
except:
     pass


Myfiles = [i for i in glob('*.xlsx')]
# print(Myfiles)
total_student = []
# 각 파일에서 학생 정보 저장.
for item in Myfiles:
    my_workbook = load_workbook(item, data_only = True)
    my_worksheet = my_workbook['Sheet1']
    my_list = []
    my_list.append(my_worksheet['A2'].value)
    my_list.append(my_worksheet['B2'].value)
    my_list.append(my_worksheet['C2'].value)
    my_list.append(my_worksheet['D2'].value)
    total_student.append(my_list)
    print(my_list)



# 전체 학생 그룹별 딕셔너리
total_student_by_group = {}
for i in range (10):
    total_student_by_group["group"+str(i+1)] = {}
print(total_student_by_group)

# 조 지정 대조 딕셔너리
assign_dict = {}
for i in range (10):
    assign_dict[i+1] = "group"+str(i+1)
print(assign_dict)

# student per cycle
student_number_tracker = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

print(total_student)
# 학생 조에 따라 분류
for item in total_student:
    group = item[2]
    group_name = assign_dict[group]
    total_student_by_group[group_name]["student_"+str(student_number_tracker[group-1]+1)] = item
    student_number_tracker[group-1] += 1

print(total_student_by_group)
print(student_number_tracker)
# my_dict = {}
# my_dict["stu_ID"] = my_worksheet['A2'].value
# my_dict["name"] = my_worksheet['B2'].value
# my_dict["group_id"] = my_worksheet['C2'].value
# my_dict["git"] = my_worksheet['D2'].value
# print(my_dict)
# print(my_worksheet['A2'].value)
# print(my_worksheet['B2'].value)
# print(my_worksheet['C2'].value)
# print(my_worksheet['D2'].value)


### 출력 영역 ###
my_writing_wb = Workbook()

# 10개의 시트 제작.
for i in range(10):
    write_ws = my_writing_wb.create_sheet("jo"+str(i+1))

# 실제 엑셀 출력 해주는 곳
for i in range(10):
    # 현재 쓰고자 하는 워크시트 로딩
    load_ws = my_writing_wb["jo"+str(i+1)]
    # 헤더 추가 (표 맨 위에 오는 정보)
    load_ws.append(["stu_num", "name", "group_id", "git"])

    # 임시로 각 조원들의 데이터 저장
    tempList = list(total_student_by_group["group"+str(i+1)].values())
    print(tempList)
    for j in range(4):
        try:
            load_ws.append(tempList[j])
        except:
            pass


my_writing_wb.remove(my_writing_wb['Sheet'])

# py 파일과 같은 폴더에 저장
my_writing_wb.save("group_python.xlsx")
